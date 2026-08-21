#!/usr/bin/env python3
"""READ-ONLY: join Price Review v3 signed Billing Price to PG effective
per-service prices. Emit verdict per row: MATCH / STALE_PG / UNMAPPED /
UNKNOWN. Diagnosis only — no fix."""
import json
import openpyxl
from collections import defaultdict

PRICE_REVIEW = "/Users/kevinfietek/Documents/Claude /Service Calendars/KitchFix_Service_Calendar_Price_Review_v3_FINAL.xlsx"
PG_JSON = "/tmp/pg-effective-prices.json"

# Load signed Price Review v3
wb = openpyxl.load_workbook(PRICE_REVIEW, data_only=True)
ws = wb["Service Price Review"]
signed_rows = []
for r in range(2, ws.max_row + 1):
    row = {
        "r": r,
        "account": ws.cell(row=r, column=1).value,
        "group":   ws.cell(row=r, column=2).value,
        "service": ws.cell(row=r, column=3).value,
        "full":    ws.cell(row=r, column=4).value,
        "billing": ws.cell(row=r, column=5).value,
        "unit":    ws.cell(row=r, column=6).value,
        "flags":   ws.cell(row=r, column=7).value,
        "notes":   ws.cell(row=r, column=8).value,
        "sfinfo":  ws.cell(row=r, column=9).value,
        "action":  ws.cell(row=r, column=10).value,
    }
    signed_rows.append(row)

# Load PG effective prices
pg_rows = json.load(open(PG_JSON))

# Normalize keys for joining
def norm_group(g):
    if g is None: return None
    s = str(g).strip()
    return s

def norm_service(s):
    if s is None: return None
    return str(s).strip()

def norm_account(a):
    if a is None: return None
    return str(a).strip()

# Build lookup indices
signed_by_key = {}
signed_billing_missing = []  # rows where billing column blank
for row in signed_rows:
    acct = norm_account(row["account"])
    grp  = norm_group(row["group"])
    svc  = norm_service(row["service"])
    if svc is None:  # Header-style rows like "Billing: Per-Meal" have svc=None
        continue
    key = (acct, grp, svc)
    if row["billing"] is None:
        signed_billing_missing.append(row)
    signed_by_key[key] = row

pg_by_key = {}
for row in pg_rows:
    acct = norm_account(row["account"])
    grp  = norm_group(row["group"])
    svc  = norm_service(row["service_name"])
    key = (acct, grp, svc)
    # Fee accounts have $0 or None in PG; keep those
    if key in pg_by_key:
        # Duplicate services under the same (acct, group, service) — flag
        prev = pg_by_key[key]
        # Prefer the row with a non-null pg_price
        if row["pg_price"] is not None and prev["pg_price"] is None:
            pg_by_key[key] = row
    else:
        pg_by_key[key] = row

# Join and verdict
verdicts = []
for key, srow in signed_by_key.items():
    acct, grp, svc = key
    pgrow = pg_by_key.get(key)
    billing = srow["billing"]
    pg_price = pgrow["pg_price"] if pgrow else None
    verdict = None
    delta = None
    reason = None
    if billing is None:
        verdict = "UNKNOWN"
        reason = "Signed Billing Price is blank"
    elif pgrow is None:
        verdict = "UNMAPPED_PG"
        reason = "Signed row has no matching PG service"
    elif pg_price is None:
        # PG has the service but no effective price
        verdict = "UNMAPPED_PG"
        reason = "PG has service but no effective price for 2026 date"
    else:
        # Compare — coerce types safely
        try:
            b = float(billing)
            p = float(pg_price)
            delta = round(p - b, 4)
            if abs(delta) < 0.01:
                verdict = "MATCH"
            else:
                verdict = "STALE_PG"
        except (TypeError, ValueError):
            verdict = "UNKNOWN"
            reason = f"Non-numeric compare: signed={billing!r} pg={pg_price!r}"
            delta = None
    verdicts.append({
        "key": key,
        "account": acct, "group": grp, "service": svc,
        "signed_billing": billing,
        "pg_price": pg_price,
        "delta": delta,
        "verdict": verdict,
        "reason": reason,
        "sfinfo": srow.get("sfinfo"),
        "flags": srow.get("flags"),
        "notes": srow.get("notes"),
        "action": srow.get("action"),
        "signed_row": srow["r"],
    })

# Also enumerate PG rows that had NO signed match
signed_keys = set(signed_by_key.keys())
pg_only = []
for key, pgrow in pg_by_key.items():
    if key not in signed_keys:
        pg_only.append({"key": key, "pg_price": pgrow["pg_price"], "pg_row": pgrow})

# Rollup
by_verdict = defaultdict(list)
for v in verdicts:
    by_verdict[v["verdict"]].append(v)

print(f"### Verdict rollup")
for k in ("MATCH", "STALE_PG", "UNMAPPED_PG", "UNKNOWN"):
    print(f"  {k}: {len(by_verdict[k])}")
print(f"  PG-only (no signed match): {len(pg_only)}")
print()

print(f"### STALE_PG fix-list ({len(by_verdict['STALE_PG'])} rows)")
for v in sorted(by_verdict["STALE_PG"], key=lambda x: (x["account"], x["group"] or "", x["service"] or "")):
    print(f"  {v['account']:12} | {v['group']:14} | {v['service']:28} | PG {v['pg_price']!r:>12} | signed {v['signed_billing']!r:>12} | delta {v['delta']!r:>10}")

print()
print(f"### UNMAPPED_PG ({len(by_verdict['UNMAPPED_PG'])} rows) — signed but no PG match")
for v in by_verdict["UNMAPPED_PG"]:
    print(f"  {v['account']} | {v['group']} | {v['service']} | signed {v['signed_billing']} | reason: {v['reason']}")

print()
print(f"### UNKNOWN ({len(by_verdict['UNKNOWN'])} rows) — signed Billing Price blank")
for v in by_verdict["UNKNOWN"]:
    print(f"  {v['account']} | {v['group']} | {v['service']} | notes: {v['notes']!r} | flags: {v['flags']!r} | action: {v['action']!r}")

print()
print(f"### PG-only ({len(pg_only)} rows) — in PG, not in signed sheet")
for v in pg_only:
    acct, grp, svc = v["key"]
    print(f"  {acct} | {grp} | {svc} | PG price {v['pg_price']}")

# Save full JSON for the write-up
with open("/tmp/price-audit-verdicts.json", "w") as f:
    json.dump({"verdicts": verdicts, "pg_only": pg_only, "signed_row_count": len(signed_rows), "pg_row_count": len(pg_rows)}, f, default=str, indent=2)
