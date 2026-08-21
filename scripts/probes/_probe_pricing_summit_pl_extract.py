#!/usr/bin/env python3
"""READ-ONLY: extract 2026 P&L revenue lines from the 11 individual-site
xlsx files under /Users/kevinfietek/Documents/KitchFix/Finance/P:L 2026/
2026 P&L Individual Sites/. Produces stdout markdown for the appendix.

- Locates revenue rows by first-column label match:
    2200 Catering Revenue
    2300 Service Charges
    2400 Meal Service (parent + Total)
    2400.1 Meal Service (Home)
    2400.2 Meal Service (Away)
    (and any other line under the "Revenue" section)
- Reads the header row (P1..P13 + Year) to know period columns.
- Emits per-file 13-period vectors + year totals + cell provenance.
- Does NOT interpret. Evidence only.
"""
import openpyxl
import os
import sys

INDIR = "/Users/kevinfietek/Documents/KitchFix/Finance/P:L 2026/2026 P&L Individual Sites"

FILE_TO_ACCOUNT = {
    "CIN - Cincinnati, OH - 2026 P&L - Clean.xlsx":  "CIN - OH",
    "CIN - Goodyear, AZ - 2026 P&L - Clean.xlsx":    "CIN - AZ",
    "CIN - Louisville, KY - 2026 P&L - Clean.xlsx":  "CIN - KY",
    "STL - Jupiter, FL - 2026 P&L - Clean.xlsx":     "STL - FL",
    "STL - St. Louis, MO - 2026 P&L - Clean.xlsx":   "STL - MO",
    "TBJ - Buffalo, NY - 2026 P&L - Clean.xlsx":     "TBJ - NY",
    "TBJ - Dunedin, FL - 2026 P&L - Clean.xlsx":     "TBJ - FL",
    "TBR - Port Charlotte, FL - 2026 P&L - Clean.xlsx": "TBR - FL",
    "TXR - H - Arlington, TX - 2026 P&L - Clean.xlsx":  "TXR - TX - H",
    "TXR - Surprise, AZ - 2026 P&L - Clean.xlsx":       "TXR - AZ",
    "TXR - V - Arlington, TX - 2026 P&L - Clean.xlsx":  "TXR - TX - V",
}

REVENUE_LABELS = [
    "2200 Catering Revenue",
    "2300 Service Charges",
    "2400 Meal Service",
    "2400.1 Meal Service",  # matches "2400.1 Meal Service (Home)" etc
    "2400.2 Meal Service",
    "Total 2400 Meal Service",
    "Total Revenue",
]


def find_period_row(ws):
    """Locate the P1..P13 header row. Returns (row_number, list-of-13-cols)."""
    for r in range(1, min(15, ws.max_row) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, min(20, ws.max_column) + 1)]
        # look for P1..P13 sequentially
        for start_col in range(1, len(row) - 12):
            slot = row[start_col: start_col + 13]
            if all(str(v).strip().upper() == f"P{i+1}" for i, v in enumerate(slot) if v):
                # column numbers are 1-indexed openpyxl; but start_col above is 0-indexed slice
                period_cols = list(range(start_col + 1, start_col + 14))
                # year column often the next non-empty
                return r, period_cols
    return None, None


def money(v):
    if v is None:
        return "-"
    if isinstance(v, (int, float)):
        if abs(v) < 0.01:
            return "0"
        return f"{v:,.0f}"
    return str(v)


def extract_file(path, account_key):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    period_row, period_cols = find_period_row(ws)
    year_col = period_cols[-1] + 1 if period_cols else None
    print(f"\n## {account_key} — `{os.path.basename(path)}` (sheet `{ws.title}`)")
    print(f"\nHeader row: `R{period_row}`; period columns `C{period_cols[0]}..C{period_cols[-1]}`; Year at `C{year_col}`.")
    print(f"\n| Line | Row | P1 | P2 | P3 | P4 | P5 | P6 | P7 | P8 | P9 | P10 | P11 | P12 | P13 | Year |")
    print(f"| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |")
    matched_any = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        # Match any revenue label
        matched = False
        for lbl in REVENUE_LABELS:
            if lbl.lower() in s.lower():
                matched = True
                break
        if not matched:
            continue
        matched_any = True
        vals = [ws.cell(row=r, column=c).value for c in period_cols]
        year = ws.cell(row=r, column=year_col).value if year_col else None
        # Only print if any value is non-zero OR the row is 2300/2200/2400.x (structural)
        any_val = year not in (None, 0, 0.0) or any(x not in (None, 0, 0.0) for x in vals)
        print(f"| {s} | R{r} | {' | '.join(money(x) for x in vals)} | {money(year)} |")
    if not matched_any:
        print("| _(no revenue-labeled rows matched)_ | | | | | | | | | | | | | | | |")


def main():
    print("# 2026 P&L per-site revenue-line extraction")
    print()
    print("READ-ONLY. Generated for the pricing summit Phase 0b brief.")
    print()
    print("Provenance per figure: filename + sheet name + row number below.")
    files = sorted(os.listdir(INDIR))
    for f in files:
        if not f.endswith(".xlsx") or f.startswith("~$"):
            continue
        acct = FILE_TO_ACCOUNT.get(f, "UNKNOWN")
        p = os.path.join(INDIR, f)
        try:
            extract_file(p, acct)
        except Exception as e:
            print(f"\n## {acct} — `{f}` — ERROR: {e}")


if __name__ == "__main__":
    main()
