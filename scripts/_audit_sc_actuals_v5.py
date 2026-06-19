#!/usr/bin/env python3
"""V5: correct audit. Service cols = row 2 has a non-metadata text label.
Data lives in the name column, not the price column."""
import openpyxl
from datetime import datetime, date
from collections import defaultdict

BASE = "/Users/kevinfietek/Documents/Claude /Service Calendars"

FILES = [
    ("CIN - AZ", "REDS AZ - Service Calendar 2026 (4).xlsx",                       "Goodyear, AZ - 2026 - Actuals", "B"),
    ("CIN - KY", "Louisville Bats Service Calendar - 2026 (2).xlsx",               "Louisville - 2026 - Actuals", "B"),
    ("STL - FL", "STL - Jupiter, FL - Service Calendar - 2026 (4).xlsx",           "Jupiter - 2026 - Actuals", "B"),
    ("TBJ - FL", "TBJ FL - Service Calendar - 2026 (4).xlsx",                      "TBJ - Actuals - 2026", "B"),
    ("TBJ - NY", "TBJ BUF - Service Calendar - 2026 (1).xlsx",                     "Buffalo - Actuals - 2026", "B"),
    ("TBR - FL", "Tampa Bay Rays Service Calendar - 2026 (3).xlsx",                "TBR-2026 - Actuals", "A"),
    ("TBR - FL B&G", "Tampa Bay Rays Service Calendar - 2026 (3).xlsx",            "B&G-2026 - Actuals", "B"),
    ("TXR - AZ", "TXR AZ - Service Calendar - 2026 (4).xlsx",                      "Actuals - 2026", "B"),
]

# Portal counts from screenshots (X entered / Y projected).
PORTAL = {
    "CIN - AZ":      {1:(31,31), 2:(28,28), 3:(31,31), 4:(30,30), 5:(31,31), 6:(14,30), 7:(0,31), 8:(0,31), 9:(0,30), 10:(0,31), 11:(0,30), 12:(0,31)},
    "STL - FL":      {1:(31,31), 2:(15,28), 3:( 1,31), 4:(27,30), 5:(30,31), 6:( 0,30), 7:(0,31), 8:(0,31), 9:(0,30), 10:(0,31), 11:(0,30), 12:(0,31)},
    "TBJ - FL":      {1:(27,31), 2:(28,28), 3:(31,31), 4:(30,30), 5:(30,31), 6:(21,30), 7:(0,31), 8:(0,31), 9:(0,30), 10:(0,31), 11:(0,30), 12:(0,20)},
    "TBR - FL":      {1:(29,31), 2:(28,28), 3:(31,31), 4:(30,30), 5:(31,31), 6:(27,30), 7:(0,31), 8:(0,31), 9:(0,30), 10:(0,31), 11:(0,30), 12:(0,29)},
    "TXR - AZ":      {1:(27,27), 2:(28,28), 3:(31,31), 4:(30,30), 5:(31,31), 6:(14,30)},
    "TBJ - NY":      {1:( 0, 0), 2:( 0, 0), 3:( 7, 9), 4:(14,30), 5:(14,31), 6:( 7,30), 7:(0,31), 8:(0,31), 9:(0,27)},
    "CIN - KY":      None,
    "TBR - FL B&G":  None,  # B&G has its own date range; portal aggregates with TBR - FL
}

MONTH_NAMES = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

METADATA_LABELS = {
    "day", "date", "period", "week", "camp", "camp name", "homestand",
    "game type", "game time", "holiday",
}
CALC_LABELS = {
    "total revenue", "total meals", "total snacks", "total bev services",
    "total charged items", "total extras", "average $/item",
}
SKIP_SERVICE_NAMES = {"blank"}


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_index(letter):
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def coerce_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def is_numeric(v):
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def find_service_cols(ws):
    """Service columns = those where row 2 has a text label that is NOT
    a metadata/calc label and NOT 'Blank'.
    Returns list of (column_index, service_name).
    """
    services = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is None or v == "":
            continue
        if not isinstance(v, str):
            continue
        norm = v.strip().lower()
        if norm in METADATA_LABELS or norm in CALC_LABELS:
            continue
        if norm in SKIP_SERVICE_NAMES:
            continue
        # If the cell value is numeric-looking (a price), skip.
        if is_numeric(v.strip()):
            continue
        services.append((c, v.strip()))
    return services


def audit_actuals(xlsx_path, tab_name, date_col_letter):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    if tab_name not in wb.sheetnames:
        return None
    ws = wb[tab_name]
    date_col = col_index(date_col_letter)
    service_cols = find_service_cols(ws)

    # Find first data row by scanning for a parseable date
    first_data_row = None
    for r in range(2, 15):
        if coerce_date(ws.cell(row=r, column=date_col).value) is not None:
            first_data_row = r
            break
    if first_data_row is None:
        return None

    dates_with_actuals = set()
    dates_with_nonzero = set()
    per_date_breakdown = {}

    for r in range(first_data_row, ws.max_row + 1):
        d = coerce_date(ws.cell(row=r, column=date_col).value)
        if d is None or d.year != 2026:
            continue
        services_with_data = []
        any_nonzero = False
        for c, name in service_cols:
            v = ws.cell(row=r, column=c).value
            if v is None or v == "":
                continue
            if is_numeric(v):
                services_with_data.append((name, float(v)))
                if float(v) > 0:
                    any_nonzero = True
        if services_with_data:
            dates_with_actuals.add(d)
            per_date_breakdown[d] = services_with_data
        if any_nonzero:
            dates_with_nonzero.add(d)

    return {
        "tab": tab_name,
        "date_col": date_col_letter,
        "first_data_row": first_data_row,
        "service_cols": service_cols,
        "dates_with_actuals": dates_with_actuals,
        "dates_with_nonzero": dates_with_nonzero,
        "per_date": per_date_breakdown,
    }


def main():
    print("=" * 84)
    print("  Service Calendar audit (V5): xlsx Actuals vs Portal year-view")
    print("  Today: 2026-06-16  |  Import: 2026-06-15")
    print()
    print("  'xlsx_any' = days with ANY numeric value in any service col (0 included)")
    print("  'xlsx_nz'  = days with at least one value > 0")
    print("  'portal'   = green 'entered' days from your screenshot")
    print("=" * 84)

    grand_totals = {"xlsx_any": 0, "xlsx_nz": 0, "portal": 0}

    for key, fname, tab, dcol in FILES:
        print()
        print(f"────── {key} ──────")
        try:
            result = audit_actuals(f"{BASE}/{fname}", tab, dcol)
        except FileNotFoundError:
            print(f"  FILE NOT FOUND: {fname}")
            continue
        if result is None:
            print(f"  could not parse tab '{tab}'")
            continue

        print(f"  tab='{result['tab']}'  date_col={result['date_col']}  "
              f"first_data_row={result['first_data_row']}  "
              f"service_cols={len(result['service_cols'])}")
        print(f"  service names: {[s[1] for s in result['service_cols']][:8]}"
              + (f" + {len(result['service_cols'])-8} more" if len(result['service_cols']) > 8 else ""))

        by_month_any = defaultdict(set)
        by_month_nz = defaultdict(set)
        for d in result["dates_with_actuals"]:
            by_month_any[d.month].add(d)
        for d in result["dates_with_nonzero"]:
            by_month_nz[d.month].add(d)

        portal = PORTAL.get(key)
        print()
        if portal:
            print(f"  {'month':<6} {'xlsx_any':>9} {'xlsx_nz':>8} {'portal':>7} {'gap_any':>8} {'gap_nz':>7}")
            print(f"  {'-'*6} {'-'*9} {'-'*8} {'-'*7} {'-'*8} {'-'*7}")
        else:
            print(f"  {'month':<6} {'xlsx_any':>9} {'xlsx_nz':>8}  (no portal screenshot for compare)")
            print(f"  {'-'*6} {'-'*9} {'-'*8}")

        t_any, t_nz, t_portal = 0, 0, 0
        gap_dates = defaultdict(list)
        for m in range(1, 13):
            xany = len(by_month_any.get(m, set()))
            xnz = len(by_month_nz.get(m, set()))
            t_any += xany
            t_nz += xnz
            pent = 0
            if portal and m in portal:
                pent, _ = portal[m]
                t_portal += pent
                gap_a = xany - pent
                gap_n = xnz - pent
                mark = ""
                if gap_n > 0:
                    mark = "  ← xlsx HAS DATA not shown"
                elif gap_n < 0 and pent > 0:
                    mark = "  ← portal counts > xlsx nonzero"
                print(f"  {MONTH_NAMES[m]:<6} {xany:>9} {xnz:>8} {pent:>7} {gap_a:+8d} {gap_n:+7d}{mark}")
                if gap_n > 0:
                    extra = sorted(by_month_nz[m])
                    gap_dates[m] = extra
            else:
                if xany > 0:
                    print(f"  {MONTH_NAMES[m]:<6} {xany:>9} {xnz:>8}")

        if portal:
            print(f"  {'-'*6} {'-'*9} {'-'*8} {'-'*7}")
            print(f"  {'TOTAL':<6} {t_any:>9} {t_nz:>8} {t_portal:>7}    nz-portal={t_nz-t_portal:+d}")
            grand_totals["xlsx_any"] += t_any
            grand_totals["xlsx_nz"] += t_nz
            grand_totals["portal"] += t_portal

        if gap_dates:
            print()
            print(f"  ⚠ DATES with xlsx actuals BEYOND portal-entered count:")
            for m, dates_list in gap_dates.items():
                print(f"    {MONTH_NAMES[m]}: xlsx_nz={len(dates_list)}, portal={portal[m][0]}, "
                      f"extra={len(dates_list)-portal[m][0]}")
                # Sample first 10 nz dates
                samples = [d.isoformat() for d in dates_list[:15]]
                print(f"      sample nz dates: {', '.join(samples)}"
                      + (f" ... (+{len(dates_list)-15} more)" if len(dates_list) > 15 else ""))

    print()
    print("=" * 84)
    print(f"  GRAND TOTALS (across compared accounts):")
    print(f"    xlsx_any (days with any value)     = {grand_totals['xlsx_any']}")
    print(f"    xlsx_nz  (days with value > 0)     = {grand_totals['xlsx_nz']}")
    print(f"    portal   (green 'entered' days)    = {grand_totals['portal']}")
    print(f"    delta nz - portal                  = {grand_totals['xlsx_nz'] - grand_totals['portal']:+d}")
    print("=" * 84)


if __name__ == "__main__":
    main()
