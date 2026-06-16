#!/usr/bin/env python3
"""
_extract_sc_xlsx.py - Service Calendar xlsx -> JSON extractor.

Reads the 11 Service Calendar spreadsheets and emits a single JSON file
that the Node seed script (_seed_sc_from_xlsx.mjs) consumes to populate
the six sc_* Postgres tables.

This is the labor side of the import pipeline. The Node script handles
the database writes and verification.

Source dir:  /Users/kevinfietek/Documents/Claude /Service Calendars/
             (NOTE the trailing space in 'Claude ')
Output JSON: scripts/_sc_extract.json

Run:  python3 scripts/_extract_sc_xlsx.py
"""

import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

SOURCE_DIR = Path("/Users/kevinfietek/Documents/Claude /Service Calendars")
OUT_PATH = Path(__file__).parent / "_sc_extract.json"

# --------------------------------------------------------------------------
# Account -> file + tab spec.
# Each account has projections + (optional) actuals tab names.
# B&G is handled inline within TBR - FL.
# --------------------------------------------------------------------------
ACCOUNTS = [
    {
        "key": "CIN - AZ",
        "file": "REDS AZ - Service Calendar 2026 (4).xlsx",
        "projections_tab": "Goodyear, AZ - Projected Number",
        "actuals_tab": "Goodyear, AZ - 2026 - Actuals",
    },
    {
        "key": "CIN - KY",
        "file": "Louisville Bats Service Calendar - 2026 (2).xlsx",
        "projections_tab": "Louisville - 2026 - Projections",
        "actuals_tab": "Louisville - 2026 - Actuals",
    },
    {
        "key": "CIN - OH",
        "file": "Cincinnati Reds MLB Service Calendar - 2026 (2).xlsx",
        "projections_tab": "Cincinnati Reds - MLB - 2026 - ",
        "actuals_tab": None,
    },
    {
        "key": "STL - FL",
        "file": "STL - Jupiter, FL - Service Calendar - 2026 (4).xlsx",
        "projections_tab": "Jupiter - 2026 - Projections",
        "actuals_tab": "Jupiter - 2026 - Actuals",
    },
    {
        "key": "STL - MO",
        "file": "St. Louis Cardinals MLB - Service Calendar - 2026 (3).xlsx",
        "projections_tab": "St. Louis MLB - 2026 - Projecti",
        "actuals_tab": "St. Louis MLB - 2026 - Actuals",
    },
    {
        "key": "TBJ - FL",
        "file": "TBJ FL - Service Calendar - 2026 (4).xlsx",
        "projections_tab": "TBJ - Projections - 2026",
        "actuals_tab": "TBJ - Actuals - 2026",
    },
    {
        "key": "TBJ - NY",
        "file": "TBJ BUF - Service Calendar - 2026 (1).xlsx",
        "projections_tab": "Buffalo - Projections - 2026",
        "actuals_tab": "Buffalo - Actuals - 2026",
    },
    {
        "key": "TBR - FL",
        "file": "Tampa Bay Rays Service Calendar - 2026 (3).xlsx",
        "projections_tab": "Projections TBR-2026",
        "actuals_tab": "TBR-2026 - Actuals",
        "bg_projections_tab": "  Projections B&G-2026",   # leading double space
        "bg_actuals_tab": "B&G-2026 - Actuals",
    },
    {
        "key": "TXR - AZ",
        "file": "TXR AZ - Service Calendar - 2026 (4).xlsx",
        "projections_tab": "Projections - 2026",
        "actuals_tab": "Actuals - 2026",
    },
    {
        "key": "TXR - TX - H",
        "file": "Texas Rangers MLB - Home - Service Calendar - 2026 (2).xlsx",
        "projections_tab": "Projections",
        "actuals_tab": None,
    },
    {
        "key": "TXR - TX - V",
        "file": "Texas Rangers MLB - Visitors - Service Calendar - 2026 (2).xlsx",
        "projections_tab": "Texas Rangers MLB V - 2026 - Pr",
        "actuals_tab": "Texas Rangers MLB V- 2026 - Act",
    },
]

# Calculated columns to deny (case-insensitive substring match on row-2 header).
DENY_HEADERS = {
    "total revenue",
    "total meals",
    "total snacks",
    "total bev services",
    "total extras",
    "total charged items",
    "average $/item",
    "total",  # standalone "Total" in clicker tab
}

# Row 2 cells where the value IS a price (these columns belong to the
# IMMEDIATELY-PREVIOUS service name column). Detected as: row-2 cell is numeric.

# Metadata column header labels (row 2). Lowercased for matching.
META_HEADERS = {
    "day", "date", "period", "week", "homestand", "camp name",
    "game type", "game time", "holiday",
    # variations
    "day of week",
}


def cell_value(ws, row, col):
    return ws.cell(row=row, column=col).value


def normalize_header(v):
    if v is None:
        return ""
    return str(v).strip()


def is_denied(label):
    if not label:
        return False
    l = label.strip().lower()
    if l in DENY_HEADERS:
        return True
    # Catch "Total" + suffix variants
    if l.startswith("total ") or l == "total":
        return True
    if l.startswith("average $"):
        return True
    return False


def is_meta_header(label):
    if not label:
        return False
    l = label.strip().lower()
    return l in META_HEADERS


def parse_date(v):
    """Return ISO date string (YYYY-MM-DD) or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Try a few formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def normalize_period(v):
    """Normalize period to integer-string (1..13). None if unparseable."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            p = int(float(v))  # 1.3 -> 1
            if 1 <= p <= 13:
                return str(p)
        except (ValueError, TypeError):
            return None
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            p = int(float(s))
            if 1 <= p <= 13:
                return str(p)
        except (ValueError, TypeError):
            return None
    return None


def normalize_game_type(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s.upper()


def normalize_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def num_value(v):
    """Return numeric float or None. Reject strings/dates."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def build_column_map(ws, max_col):
    """
    Build a column map by scanning row 1 (group headers) and row 2
    (service name + price interleaved).

    Returns:
      meta_cols:    dict label_lower -> col_index
      services:     list of dicts:
                      { group_name, service_name, price (or None), name_col, price_col, sort_in_group }
      denied_cols:  set of col_index (calculated columns)
    """
    # First, scan row 1 to determine group ranges.
    # Group header lives in the leftmost cell of its range; merged cells will only
    # have value on the topleft. openpyxl returns None for the others.
    group_at_col = {}
    last_group = None
    group_first_col = {}  # group_name -> leftmost col where it appeared
    for c in range(1, max_col + 1):
        v = cell_value(ws, 1, c)
        v_norm = normalize_header(v)
        if v_norm:
            last_group = v_norm
            if last_group not in group_first_col:
                group_first_col[last_group] = c
        group_at_col[c] = last_group  # may be None for cells before any group

    # Scan row 2 for service names + prices and metadata.
    meta_cols = {}
    services = []
    denied_cols = set()
    pending_name = None  # (col, name, group)

    # Per-group counters for sort_in_group
    group_service_count = {}

    c = 1
    while c <= max_col:
        v = cell_value(ws, 2, c)
        label = normalize_header(v)
        label_l = label.lower()
        group = group_at_col.get(c)

        # Calculated column block?
        if is_denied(label):
            denied_cols.add(c)
            # Reset pending: a denied col cannot be a price col.
            pending_name = None
            c += 1
            continue

        # Metadata?
        if is_meta_header(label):
            meta_cols[label_l] = c
            pending_name = None
            c += 1
            continue

        # Group is TOTALS? skip
        if group and group.strip().upper() == "TOTALS":
            pending_name = None
            c += 1
            continue

        # Numeric row-2 value? It's a price for the previous name col.
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if pending_name is not None:
                pname_col, pname, pgroup = pending_name
                # apply price
                services[-1]["price"] = float(v)
                services[-1]["price_col"] = c
                pending_name = None
            else:
                # Stray number. Ignore.
                pass
            c += 1
            continue

        # String header that is not a meta label and not denied.
        # STL-FL Fun Money pattern: row 1 has group "Fun Money" at col AE,
        # row 2 at AE is empty, and row 2 at AF is the price 25000. There's
        # no row-2 service-name cell. Detect by: empty row-2 here, group is
        # set, group's leftmost col is this one, and next col is numeric.
        if not label:
            next_val = cell_value(ws, 2, c + 1) if c + 1 <= max_col else None
            if (
                group
                and group_first_col.get(group) == c
                and group_service_count.get(group, 0) == 0
                and isinstance(next_val, (int, float))
                and not isinstance(next_val, bool)
            ):
                sort_in_group = group_service_count.get(group, 0)
                group_service_count[group] = sort_in_group + 1
                services.append({
                    "group_name": group,
                    "service_name": f"{group} allocation",
                    "price": float(next_val),
                    "name_col": c + 1,
                    "price_col": c + 1,
                    "sort_in_group": sort_in_group,
                })
                pending_name = None
                c += 2
                continue
            pending_name = None
            c += 1
            continue

        # It's a service name (potentially "Blank").
        # The next column over should hold the price (numeric in row 2).
        next_val = cell_value(ws, 2, c + 1) if c + 1 <= max_col else None
        price = None
        price_col = None
        if isinstance(next_val, (int, float)) and not isinstance(next_val, bool):
            price = float(next_val)
            price_col = c + 1

        sort_in_group = group_service_count.get(group, 0)
        group_service_count[group] = sort_in_group + 1

        services.append({
            "group_name": group or "",
            "service_name": label,
            "price": price,
            "name_col": c,
            "price_col": price_col,
            "sort_in_group": sort_in_group,
        })
        pending_name = (c, label, group) if price is None else None
        # If we set price from c+1, skip that column.
        if price_col is not None:
            c += 2
        else:
            c += 1

    return meta_cols, services, denied_cols, group_first_col


def extract_tab(ws, max_col=None):
    """
    Read a sheet and return:
      column_map: { meta_cols, services, denied_cols, group_first_col }
      rows:        list of { date: 'YYYY-MM-DD', meta: {...}, values: { name_col: number } }
    """
    if max_col is None:
        max_col = ws.max_column
    meta_cols, services, denied_cols, group_first_col = build_column_map(ws, max_col)

    # Find the date column (case-insensitive 'date')
    date_col = meta_cols.get("date")
    if date_col is None:
        return {
            "meta_cols": meta_cols,
            "services": services,
            "denied_cols": sorted(denied_cols),
            "group_first_col": group_first_col,
            "rows": [],
            "skipped_invalid_dates": 0,
        }

    rows = []
    skipped = 0
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        d_raw = cell_value(ws, r, date_col)
        d_iso = parse_date(d_raw)
        if d_iso is None:
            # Stray row (Proposed Increase etc.).
            # Only count as skipped if there's any data in the row.
            any_val = False
            for c in range(1, max_col + 1):
                if cell_value(ws, r, c) is not None:
                    any_val = True
                    break
            if any_val:
                skipped += 1
            continue

        meta = {}
        for label, col in meta_cols.items():
            if label == "date":
                continue
            meta[label] = cell_value(ws, r, col)

        # Service values keyed by name_col
        vals = {}
        for s in services:
            v = cell_value(ws, r, s["name_col"])
            n = num_value(v)
            if n is not None:
                vals[str(s["name_col"])] = n
        rows.append({"date": d_iso, "meta": meta, "values": vals})

    return {
        "meta_cols": {k: v for k, v in meta_cols.items()},
        "services": services,
        "denied_cols": sorted(denied_cols),
        "group_first_col": group_first_col,
        "rows": rows,
        "skipped_invalid_dates": skipped,
    }


def extract_bg_tab(ws):
    """
    B&G tabs have a simpler layout:
      col A: Day
      col B: Date
      col C: Period
      col D: Lunch (service)
      col E: 6.5 (price for Lunch in row 2)
      col F: Total Revenue (SKIP)
      col G: Holiday
    """
    # We synthesize a single service: B&G Lunch
    max_col = ws.max_column
    max_row = ws.max_row

    # Locate by header labels in row 2 (more robust than column letters).
    # B&G has its row-1 label "Boys & Girls Club" or similar.
    meta_cols = {}
    name_col = None
    price = 6.50

    # Scan row 2 for Date, Period, Lunch, Holiday (skip Total Revenue)
    for c in range(1, max_col + 1):
        v = cell_value(ws, 2, c)
        label = normalize_header(v)
        l = label.lower()
        if l == "date":
            meta_cols["date"] = c
        elif l == "period":
            meta_cols["period"] = c
        elif l == "day":
            meta_cols["day"] = c
        elif l == "week":
            meta_cols["week"] = c
        elif l == "holiday":
            meta_cols["holiday"] = c
        elif l == "lunch":
            name_col = c
            # Confirm next-col price
            nv = cell_value(ws, 2, c + 1) if c + 1 <= max_col else None
            if isinstance(nv, (int, float)) and not isinstance(nv, bool):
                price = float(nv)

    services = [{
        "group_name": "Boys & Girls Club",
        "service_name": "B&G Lunch",
        "price": price,
        "name_col": name_col,
        "price_col": (name_col + 1) if name_col else None,
        "sort_in_group": 0,
    }]

    date_col = meta_cols.get("date")
    rows = []
    skipped = 0
    if date_col is None:
        return {
            "meta_cols": meta_cols,
            "services": services,
            "denied_cols": [],
            "group_first_col": {"Boys & Girls Club": name_col or 0},
            "rows": [],
            "skipped_invalid_dates": 0,
        }

    for r in range(3, max_row + 1):
        d_raw = cell_value(ws, r, date_col)
        d_iso = parse_date(d_raw)
        if d_iso is None:
            any_val = False
            for c in range(1, max_col + 1):
                if cell_value(ws, r, c) is not None:
                    any_val = True
                    break
            if any_val:
                skipped += 1
            continue

        meta = {}
        for label, col in meta_cols.items():
            if label == "date":
                continue
            meta[label] = cell_value(ws, r, col)

        vals = {}
        if name_col is not None:
            v = cell_value(ws, r, name_col)
            n = num_value(v)
            if n is not None:
                vals[str(name_col)] = n
        rows.append({"date": d_iso, "meta": meta, "values": vals})

    return {
        "meta_cols": meta_cols,
        "services": services,
        "denied_cols": [],
        "group_first_col": {"Boys & Girls Club": name_col or 0},
        "rows": rows,
        "skipped_invalid_dates": skipped,
    }


def make_serializable(obj):
    """Recursively convert datetime objects to ISO strings."""
    if isinstance(obj, dict):
        return {k: make_serializable(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_serializable(v) for v in obj]
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    return obj


def main():
    print(f"Source dir: {SOURCE_DIR}")
    print(f"Output: {OUT_PATH}")
    if not SOURCE_DIR.exists():
        print(f"ERROR: source dir not found: {SOURCE_DIR}", file=sys.stderr)
        sys.exit(1)

    output = {"accounts": {}, "errors": []}

    for acct in ACCOUNTS:
        key = acct["key"]
        fpath = SOURCE_DIR / acct["file"]
        if not fpath.exists():
            msg = f"MISSING FILE for {key}: {fpath}"
            print(msg, file=sys.stderr)
            output["errors"].append(msg)
            output["accounts"][key] = None
            continue
        print(f"\n=== {key} === ({acct['file']})")
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=False)
        except Exception as e:
            msg = f"FAILED to open {key}: {e}"
            print(msg, file=sys.stderr)
            output["errors"].append(msg)
            output["accounts"][key] = None
            continue

        account_payload = {
            "projections": None,
            "actuals": None,
            "bg_projections": None,
            "bg_actuals": None,
        }

        # Projections tab
        pname = acct["projections_tab"]
        if pname in wb.sheetnames:
            ws = wb[pname]
            account_payload["projections"] = extract_tab(ws)
            p = account_payload["projections"]
            print(f"  projections '{pname}': {len(p['services'])} services, {len(p['rows'])} rows, skipped {p['skipped_invalid_dates']}")
        else:
            msg = f"{key} projections tab not found: '{pname}'. Sheets: {wb.sheetnames}"
            print(msg, file=sys.stderr)
            output["errors"].append(msg)

        # Actuals tab
        aname = acct.get("actuals_tab")
        if aname:
            if aname in wb.sheetnames:
                ws = wb[aname]
                account_payload["actuals"] = extract_tab(ws)
                a = account_payload["actuals"]
                print(f"  actuals     '{aname}': {len(a['services'])} services, {len(a['rows'])} rows, skipped {a['skipped_invalid_dates']}")
            else:
                msg = f"{key} actuals tab not found: '{aname}'. Sheets: {wb.sheetnames}"
                print(msg, file=sys.stderr)
                output["errors"].append(msg)

        # B&G tabs (TBR - FL only)
        bg_p = acct.get("bg_projections_tab")
        if bg_p:
            if bg_p in wb.sheetnames:
                ws = wb[bg_p]
                account_payload["bg_projections"] = extract_bg_tab(ws)
                b = account_payload["bg_projections"]
                print(f"  bg_proj     '{bg_p}': {len(b['rows'])} rows")
            else:
                msg = f"{key} B&G projections tab not found: '{bg_p}'. Sheets: {wb.sheetnames}"
                print(msg, file=sys.stderr)
                output["errors"].append(msg)

        bg_a = acct.get("bg_actuals_tab")
        if bg_a:
            if bg_a in wb.sheetnames:
                ws = wb[bg_a]
                account_payload["bg_actuals"] = extract_bg_tab(ws)
                b = account_payload["bg_actuals"]
                print(f"  bg_actuals  '{bg_a}': {len(b['rows'])} rows")
            else:
                msg = f"{key} B&G actuals tab not found: '{bg_a}'. Sheets: {wb.sheetnames}"
                print(msg, file=sys.stderr)
                output["errors"].append(msg)

        wb.close()
        output["accounts"][key] = account_payload

    # Convert any non-JSON-serializable values (datetimes etc.) to strings.
    output = make_serializable(output)

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=1, default=str)
    print(f"\nWrote {OUT_PATH} ({os.path.getsize(OUT_PATH)} bytes)")


if __name__ == "__main__":
    main()
