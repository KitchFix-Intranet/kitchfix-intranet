#!/usr/bin/env python3
"""Helper: read each PDC/MiLB account's Actuals tab and dump
per-(date, group, service) entries to JSON.

Usage: python3 _audit_sc_xlsx_dump.py /tmp/out.json
"""
import json
import sys
import openpyxl
from datetime import datetime, date

BASE = "/Users/kevinfietek/Documents/Claude /Service Calendars"

# (account_key, xlsx filename, actuals tab name, date column letter,
#  bg_account_key_or_None for files where multiple tabs feed the same account)
FILES = [
    ("CIN - AZ",     "REDS AZ - Service Calendar 2026 (4).xlsx",           "Goodyear, AZ - 2026 - Actuals", "B", None),
    ("CIN - KY",     "Louisville Bats Service Calendar - 2026 (2).xlsx",   "Louisville - 2026 - Actuals",   "B", None),
    ("STL - FL",     "STL - Jupiter, FL - Service Calendar - 2026 (4).xlsx","Jupiter - 2026 - Actuals",     "B", None),
    ("TBJ - FL",     "TBJ FL - Service Calendar - 2026 (4).xlsx",          "TBJ - Actuals - 2026",          "B", None),
    ("TBJ - NY",     "TBJ BUF - Service Calendar - 2026 (1).xlsx",         "Buffalo - Actuals - 2026",      "B", None),
    ("TBR - FL",     "Tampa Bay Rays Service Calendar - 2026 (3).xlsx",    "TBR-2026 - Actuals",            "A", None),
    ("TBR - FL",     "Tampa Bay Rays Service Calendar - 2026 (3).xlsx",    "B&G-2026 - Actuals",            "B", "Boys & Girls Club"),
    ("TXR - AZ",     "TXR AZ - Service Calendar - 2026 (4).xlsx",          "Actuals - 2026",                "B", None),
]

METADATA_LABELS = {"day","date","period","week","camp","camp name","homestand",
                   "game type","game time","holiday"}
CALC_LABELS = {"total revenue","total meals","total snacks","total bev services",
               "total charged items","total extras","average $/item"}
SKIP_NAMES = {"blank"}

# Service-name canonicalization mirroring the seed's ACTUALS_NAME_REMAP +
# B&G hardcoding. The audit's xlsx side must produce the canonical names
# stored in PG, not the raw row-2 labels.
SERVICE_NAME_REMAP = {
    # STL - FL Palm Beach Cardinals: actuals col U reads "Breakfast",
    # seed renames to "Arrival" per ACTUALS_NAME_REMAP
    ("STL - FL", "Palm Beach Cardinals", "Breakfast"): "Arrival",
}
BG_GROUP_NAME = "Boys & Girls Club"
BG_SERVICE_NAME = "B&G Lunch"


def col_index(letter):
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def coerce_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def is_numeric(v):
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.replace("$", "").replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def to_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        return float(v.replace("$", "").replace(",", ""))
    return None


def derive_group_map(ws):
    """Walk row 1 left-to-right and assign each column to the most-recent
    non-empty row-1 value. Skip TOTALS group."""
    current = None
    col_to_group = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() != "":
            current = str(v).strip()
        col_to_group[c] = current
    return col_to_group


def find_service_cols(ws):
    """Service columns = row 2 has a text label that's not metadata/calc/blank."""
    services = []  # (col, name)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is None or v == "":
            continue
        if not isinstance(v, str):
            continue
        norm = v.strip().lower()
        if norm in METADATA_LABELS or norm in CALC_LABELS or norm in SKIP_NAMES:
            continue
        if is_numeric(v.strip()):
            continue
        services.append((c, v.strip()))
    return services


def find_first_data_row(ws, date_col):
    for r in range(2, 15):
        if coerce_date(ws.cell(row=r, column=date_col).value) is not None:
            return r
    return None


def dump_tab(xlsx_path, tab_name, date_col_letter, bg_group_override):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    if tab_name not in wb.sheetnames:
        return None
    ws = wb[tab_name]
    date_col = col_index(date_col_letter)
    service_cols = find_service_cols(ws)
    if not service_cols:
        return None
    first_row = find_first_data_row(ws, date_col)
    if first_row is None:
        return None
    col_to_group = derive_group_map(ws)

    rows = []
    for r in range(first_row, ws.max_row + 1):
        d = coerce_date(ws.cell(row=r, column=date_col).value)
        if d is None or d.year != 2026:
            continue
        entries = []
        for c, svc_name in service_cols:
            v = ws.cell(row=r, column=c).value
            if v is None or v == "":
                continue
            if not is_numeric(v):
                continue
            num = to_number(v)
            # Filter out columns that are calc-totals (some files have
            # them positioned within service area). Skip if the col
            # belongs to TOTALS group.
            group = bg_group_override if bg_group_override else (col_to_group.get(c) or "?")
            if "totals" in group.lower():
                continue
            # Apply canonicalization to match PG storage names.
            # B&G tab is identified by the override; force its service name.
            if bg_group_override == BG_GROUP_NAME:
                canon_service = BG_SERVICE_NAME
            else:
                canon_service = svc_name
            entries.append({
                "col": c,
                "group": group,
                "service": canon_service,
                "value": int(num) if num == int(num) else num,
            })
        if entries:
            rows.append({"date": d.isoformat(), "entries": entries})
    return rows


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else "/tmp/_sc_audit_xlsx_dump.json"
    accumulated = {}  # account_key -> { rows: [] }
    for account_key, fname, tab, dcol, bg in FILES:
        path = f"{BASE}/{fname}"
        try:
            rows = dump_tab(path, tab, dcol, bg)
        except FileNotFoundError:
            print(f"  WARN: {path} not found", file=sys.stderr)
            continue
        if rows is None:
            print(f"  WARN: {account_key} / {tab} produced no rows", file=sys.stderr)
            continue

        # Apply per-account service-name REMAPs (mirrors seed's ACTUALS_NAME_REMAP).
        for row in rows:
            for e in row["entries"]:
                remap_key = (account_key, e["group"], e["service"])
                if remap_key in SERVICE_NAME_REMAP:
                    e["service"] = SERVICE_NAME_REMAP[remap_key]

        if account_key not in accumulated:
            accumulated[account_key] = {"rows": []}
        # If this account already has rows (e.g. TBR - FL gets two tabs),
        # merge by date.
        existing_by_date = {row["date"]: row for row in accumulated[account_key]["rows"]}
        for row in rows:
            if row["date"] in existing_by_date:
                existing_by_date[row["date"]]["entries"].extend(row["entries"])
            else:
                accumulated[account_key]["rows"].append(row)

    with open(out_path, "w") as f:
        json.dump(accumulated, f, indent=1)
    print(f"Wrote {out_path} ({sum(len(d['rows']) for d in accumulated.values())} total date-rows across {len(accumulated)} accounts)")


if __name__ == "__main__":
    main()
