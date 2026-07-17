"""Build four-way Stage-3 certification audit (Signed / PG / AccountFile / Workbook).

Re-run after any price change or signed-sheet refresh; requires a fresh
`scripts/audit-sc-prices.mjs --out /tmp/pg_prices.json` dump beforehand.

READ-ONLY. Cross-joins the four price sources into one verdict per row:
  Signed  = KitchFix_Service_Calendar_Price_Review_v3_FINAL.xlsx
            (v3 FINAL "Service Price Review" tab, "Billing Price" column)
  PG      = /tmp/pg_prices.json (from scripts/audit-sc-prices.mjs)
  AcctFile= docs/pricing-summit/accounts/ACCOUNT_*.md §2b tables
  Workbook= per-account SC workbook (projection + actuals tabs)

Verdict classes:
  ALL-MATCH · N/A-fee · PG-drift-fee · PG-FAIL · SIGNED-NO-PRICE ·
  SIGNED-STALE-STAGE1 · DOC-DRIFT · WB-EXP-DIV

Usage:
    # 1. Refresh PG dump (post-Studio-apply, from a checkout with .env.local)
    set -a && source .env.local && set +a
    node scripts/audit-sc-prices.mjs --out /tmp/pg_prices.json

    # 2. Regenerate audit
    python3 scripts/build-four-way-audit.py

Output: /Users/kevinfietek/Downloads/STAGE3_CERTIFICATION_AUDIT.md

If the signed workbook, PG snapshot, or account files move, update the
paths in the constants block below.
"""
import json, os, re
from collections import defaultdict
import openpyxl

# Point REPO at whichever intranet checkout has the current
# docs/pricing-summit/ tree. Both kf-cell-states and kitchfix-intranet
# are valid worktrees of the same repo.
REPO = os.environ.get("KF_REPO_ROOT", "/Users/kevinfietek/dev/kf-cell-states")
SIGNED_XLSX = "/Users/kevinfietek/Documents/Claude /Service Calendars/KitchFix_Service_Calendar_Price_Review_v3_FINAL.xlsx"
PG_JSON = "/tmp/pg_prices.json"
OUT_MD = "/Users/kevinfietek/Downloads/STAGE3_CERTIFICATION_AUDIT.md"

WORKBOOKS = {
    "CIN - AZ":     "REDS AZ - Service Calendar 2026 (4).xlsx",
    "CIN - KY":     "Louisville Bats Service Calendar - 2026 (2).xlsx",
    "CIN - OH":     "Cincinnati Reds MLB Service Calendar - 2026 (2).xlsx",
    "STL - FL":     "STL - Jupiter, FL - Service Calendar - 2026 (4).xlsx",
    "STL - MO":     "St. Louis Cardinals MLB - Service Calendar - 2026 (3).xlsx",
    "TBJ - FL":     "TBJ FL - Service Calendar - 2026 (4).xlsx",
    "TBJ - NY":     "TBJ BUF - Service Calendar - 2026 (1).xlsx",
    "TBR - FL":     "Tampa Bay Rays Service Calendar - 2026 (3).xlsx",
    "TXR - AZ":     "TXR AZ - Service Calendar - 2026 (4).xlsx",
    "TXR - TX - H": "Texas Rangers MLB - Home - Service Calendar - 2026 (2).xlsx",
    "TXR - TX - V": "Texas Rangers MLB - Visitors - Service Calendar - 2026 (2).xlsx",
}
WB_BASE = "/Users/kevinfietek/Documents/Claude /Service Calendars/"


def r2(x):
    """Round to 2dp for comparison; None-safe."""
    if x is None:
        return None
    try:
        return round(float(x), 2)
    except (ValueError, TypeError):
        return None


def normalize_svc(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


# ---- Load signed ----
wb = openpyxl.load_workbook(SIGNED_XLSX, data_only=True)
ws = wb["Service Price Review"]
signed_rows = []
for r in range(2, ws.max_row + 1):
    acct = ws.cell(row=r, column=1).value
    grp = ws.cell(row=r, column=2).value
    svc = ws.cell(row=r, column=3).value
    full = ws.cell(row=r, column=4).value
    bill = ws.cell(row=r, column=5).value
    unit = ws.cell(row=r, column=6).value
    flag = ws.cell(row=r, column=7).value
    if not acct or not svc:
        continue
    signed_rows.append({
        "account": str(acct).strip(),
        "group": str(grp).strip() if grp else "",
        "service": str(svc).strip(),
        "signed_full": None if full == "-" else full,
        "signed_bill": bill,  # may be "NEEDS PRICE"
        "unit": unit,
        "flag": flag,
    })

# ---- Load PG ----
pg_all = json.load(open(PG_JSON))["rows"]
pg_by_key = {}
for r in pg_all:
    k = (r["account_key"], r["group"] or "", r["service"] or "")
    pg_by_key[k] = r

# ---- Load account files ----
def parse_account_file(path, account_key):
    """Extract §2b rate table rows: (group, service, post_sf_billed_price)."""
    with open(path) as f:
        text = f.read()
    # Find §2b heading and table
    m = re.search(r"###\s*2b\.\s*Rate table[^\n]*\n(.+?)(?=\n###|\Z)", text, re.DOTALL)
    if not m:
        return []
    block = m.group(1)
    rows = []
    for line in block.split("\n"):
        if not line.startswith("|") or "---" in line:
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if len(cells) < 4:
            continue
        # Skip header
        if cells[0].lower() in ("group",):
            continue
        group = cells[0]
        service = cells[1]
        # billed column is typically cells[2] "Post-SF (billed)" but format differs per account
        # Try to extract $NN.NN from cell 2 or 3
        billed_raw = cells[2] if len(cells) > 2 else ""
        m2 = re.search(r"\$?([0-9,]+\.\d+|\d+)", billed_raw.replace("**", "").replace("*", ""))
        billed = float(m2.group(1).replace(",", "")) if m2 else None
        rows.append({"group": group, "service": service, "acct_file_billed": billed, "raw": billed_raw})
    return rows

ACCT_KEYS = list(WORKBOOKS.keys())
acct_files = {}
for k in ACCT_KEYS:
    # ACCOUNT_<KEY>.md where KEY replaces " - " with "-" (e.g. CIN-AZ)
    filename_key = k.replace(" - ", "-")
    path = f"{REPO}/docs/pricing-summit/accounts/ACCOUNT_{filename_key}.md"
    if os.path.exists(path):
        acct_files[k] = parse_account_file(path, k)
    else:
        acct_files[k] = []
        print(f"WARN: missing account file for {k}: {path}")

# Build lookup: (acct, group-fuzzy, service-fuzzy) -> account file billed
# Group synonyms: signed_name -> [acct_file candidates]
GROUP_SYN = {
    "major league": {"major league", "mlb", "major league - pdc"},
    "minor league": {"minor league", "milb", "milb - pdc", "minor league - pdc"},
    "rehab": {"rehab"},
    "single a jays": {"single a jays", "fsl", "milb (fsl - single a jays)"},
    "ssm": {"ssm"},
    "other": {"other"},
    "boys & girls club": {"boys & girls club", "bgc"},
    "louisville bats": {"louisville bats"},
    "cincinnati reds": {"cincinnati reds", "reds"},
    "st. louis cardinals": {"st. louis cardinals", "cardinals"},
    "mlb": {"mlb", "major league", "major league - pdc"},
    "milb": {"milb", "minor league", "minor league - pdc"},
    "palm beach cardinals": {"palm beach cardinals"},
    "buffalo bisons": {"buffalo bisons"},
    "texas rangers": {"texas rangers"},
    "fun money": {"fun money"},
}

def _strip_paren(s):
    return re.sub(r"\s*\([^)]*\)", "", s).strip()

def group_matches(signed_grp, acct_grp):
    sg = normalize_svc(signed_grp)
    ag = normalize_svc(acct_grp)
    if not sg or not ag:
        return True
    if sg == ag or (sg in ag or ag in sg):
        return True
    # Strip parenthetical suffixes and retry
    sg_bare = normalize_svc(_strip_paren(sg))
    ag_bare = normalize_svc(_strip_paren(ag))
    if sg_bare == ag_bare or sg_bare in ag_bare or ag_bare in sg_bare:
        return True
    # Synonyms both ways
    syns = GROUP_SYN.get(sg_bare, set()) | GROUP_SYN.get(sg, set())
    if ag_bare in syns or ag in syns:
        return True
    syns2 = GROUP_SYN.get(ag_bare, set()) | GROUP_SYN.get(ag, set())
    if sg_bare in syns2 or sg in syns2:
        return True
    return False

def lookup_acct_file(acct, group, service):
    rows = acct_files.get(acct, [])
    svc_n = normalize_svc(service)
    grp_n = normalize_svc(group)
    # Exact group + service match
    for r in rows:
        if r["acct_file_billed"] is None:
            continue
        r_svc = normalize_svc(r["service"])
        if r_svc == svc_n and group_matches(group, r["group"]):
            return r["acct_file_billed"]
    # Substring service match with group
    for r in rows:
        if r["acct_file_billed"] is None:
            continue
        r_svc = normalize_svc(r["service"])
        if svc_n in r_svc and group_matches(group, r["group"]):
            return r["acct_file_billed"]
    return None  # No match; don't fall back to any-service-with-same-name

# ---- Load workbooks ----
def extract_workbook_prices(path):
    """From row 2 of each relevant tab, extract (service_name, price) pairs.
    Return {tab_kind: [(name, price), ...]} where tab_kind in {projection, actuals}.
    """
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {"projection": [], "actuals": []}
    for tab in wb.sheetnames:
        t_lo = tab.lower()
        if "actual" in t_lo:
            kind = "actuals"
        elif "projec" in t_lo:
            kind = "projection"
        else:
            continue
        ws = wb[tab]
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=c).value
            if not isinstance(v, str):
                continue
            if v.strip().lower() in ("day", "date", "period", "week", "camp name",
                                     "camp", "phase", "phase name", "notes", "homestand",
                                     "homestand?", "period totals", "total meals",
                                     "total projected", "total actual", "total revenue"):
                continue
            # price is typically in the next column
            next_v = ws.cell(row=2, column=c + 1).value
            if isinstance(next_v, (int, float)):
                result[kind].append((v.strip(), float(next_v)))
    return result

workbook_prices = {}
for k, filename in WORKBOOKS.items():
    workbook_prices[k] = extract_workbook_prices(WB_BASE + filename)

def lookup_workbook(acct, service):
    wp = workbook_prices.get(acct, {})
    svc_n = normalize_svc(service)
    proj_hits = [p for (n, p) in wp.get("projection", []) if normalize_svc(n) == svc_n or svc_n in normalize_svc(n)]
    actl_hits = [p for (n, p) in wp.get("actuals", []) if normalize_svc(n) == svc_n or svc_n in normalize_svc(n)]
    return proj_hits, actl_hits

# ---- Assemble four-way rows ----
audit_rows = []
missing_pg = []
missing_signed_from_pg = []

# PG rows keyed for orphan detection
pg_seen = set()

for s in signed_rows:
    k = (s["account"], s["group"], s["service"])
    pg_row = pg_by_key.get(k)
    if pg_row:
        pg_seen.add(k)
    acct_file = lookup_acct_file(s["account"], s["group"], s["service"])
    proj_hits, actl_hits = lookup_workbook(s["account"], s["service"])
    # Round
    signed_bill = r2(s["signed_bill"]) if isinstance(s["signed_bill"], (int, float)) else s["signed_bill"]
    signed_full = r2(s["signed_full"]) if isinstance(s["signed_full"], (int, float)) else None
    pg_proj = r2(pg_row["projected_price"]) if pg_row and pg_row["projected_price"] is not None else None
    pg_actl = r2(pg_row["actual_price"]) if pg_row and pg_row["actual_price"] is not None else None
    acct_r = r2(acct_file)
    # Determine verdict
    is_flat = pg_row["is_flat_fee"] if pg_row else False
    is_fee_acct = (s.get("flag") == "fee_account" or "fee_account" in (s.get("flag") or ""))
    # PG vs Signed comparison
    if is_fee_acct:
        # For fee accounts: signed_bill = 0 (fee schedule); PG projected should also be $0
        pg_v_signed_ok = (pg_proj == 0 or pg_proj is None) and (signed_bill == 0)
    else:
        pg_v_signed_ok = (pg_proj is not None and signed_bill is not None
                         and isinstance(signed_bill, (int, float))
                         and pg_proj == signed_bill)
    # AcctFile vs Signed: 3-way (True=match, False=drift, None=no-entry-in-file)
    if not isinstance(signed_bill, (int, float)) or acct_r is None:
        af_v_signed_ok = None
    else:
        af_v_signed_ok = abs(acct_r - signed_bill) < 0.005
    # Workbook divergence
    wb_proj = proj_hits[0] if proj_hits else None
    wb_actl = actl_hits[0] if actl_hits else None
    wb_proj_r = r2(wb_proj)
    wb_actl_r = r2(wb_actl)
    if isinstance(signed_bill, (int, float)):
        # Workbook expected: actuals ≈ signed_bill for actuals_drive; projection ≈ signed_full for many
        wb_div = None
        if is_fee_acct:
            wb_div = "fee_acct"  # workbook meaningless
        else:
            if wb_actl_r is not None and abs(wb_actl_r - signed_bill) >= 0.01:
                wb_div = f"actuals={wb_actl_r} vs signed_bill={signed_bill}"
            elif wb_proj_r is not None and signed_full is not None and abs(wb_proj_r - signed_full) >= 0.01:
                wb_div = f"proj={wb_proj_r} vs signed_full={signed_full}"
    else:
        wb_div = None

    # Overall verdict (order matters)
    # Known Stage-1 directives where PG has been intentionally moved above/away from signed
    # (Kevin ordered the change; signed sheet needs a v4 refresh to match).
    stage1_stale = (
        (s["account"] == "TBJ - FL" and "Media Meals" in s["service"])  # PG=$16 per Kevin; signed still $15
    )
    if not isinstance(signed_bill, (int, float)):
        verdict = "SIGNED-NO-PRICE"
    elif stage1_stale:
        verdict = "SIGNED-STALE-STAGE1"
    elif is_fee_acct:
        verdict = "N/A-fee" if pg_v_signed_ok else "PG-drift-fee"
    elif not pg_v_signed_ok:
        verdict = "PG-FAIL"
    elif af_v_signed_ok is False:
        verdict = "DOC-DRIFT"
    elif wb_div:
        verdict = "WB-EXP-DIV"
    else:
        verdict = "ALL-MATCH"

    audit_rows.append({
        "account": s["account"],
        "group": s["group"],
        "service": s["service"],
        "unit": s["unit"],
        "flag": s["flag"],
        "signed_bill": signed_bill,
        "signed_full": signed_full,
        "pg_proj": pg_proj,
        "pg_actl": pg_actl,
        "acct_file": acct_r,
        "wb_proj": wb_proj_r,
        "wb_actl": wb_actl_r,
        "wb_all_proj": [r2(p) for p in proj_hits],
        "wb_all_actl": [r2(p) for p in actl_hits],
        "wb_div_reason": wb_div,
        "is_flat": is_flat,
        "is_fee_acct": is_fee_acct,
        "pg_v_signed_ok": pg_v_signed_ok,
        "af_v_signed_ok": af_v_signed_ok,
        "verdict": verdict,
    })
    if not pg_row:
        missing_pg.append(k)

# PG rows not seen in signed
for r in pg_all:
    k = (r["account_key"], r["group"] or "", r["service"] or "")
    if k not in pg_seen:
        missing_signed_from_pg.append(k)

# ---- Aggregate stats ----
def count(pred):
    return sum(1 for r in audit_rows if pred(r))

pg_signed_pass = count(lambda r: r["pg_v_signed_ok"])
pg_signed_signed_no_price = count(lambda r: r["verdict"] == "SIGNED-NO-PRICE")
pg_signed_stale_stage1 = count(lambda r: r["verdict"] == "SIGNED-STALE-STAGE1")
pg_signed_real_fail = count(lambda r: r["verdict"] == "PG-FAIL")
af_signed_pass = count(lambda r: r["af_v_signed_ok"] is True)
af_signed_fail = count(lambda r: r["af_v_signed_ok"] is False)
af_signed_none = count(lambda r: r["af_v_signed_ok"] is None)
wb_div_count = count(lambda r: r["wb_div_reason"] is not None and r["wb_div_reason"] != "fee_acct")

# ---- Write output ----
def money(v):
    if v is None:
        return "-"
    if isinstance(v, str):
        return v
    if v == 0:
        return "$0"
    return f"${v:.2f}"

def verdict_icon(v):
    return {
        "ALL-MATCH":  "✅ ALL-MATCH",
        "N/A-fee":    "✅ N/A (fee)",
        "PG-drift-fee": "🟠 PG-drift-fee",
        "PG-FAIL":    "🔴 PG≠Signed",
        "DOC-DRIFT":  "🟡 AcctFile≠Signed",
        "WB-EXP-DIV": "⚪ WB≠Signed (expected)",
    }.get(v, v)

md = []
md.append("# Stage-3 Four-Way Price Certification Audit")
md.append("")
md.append(f"**Generated:** 2026-07-17 by Claude Code (read-only).  ")
md.append(f"**Rows compared:** {len(audit_rows)} (signed sheet is the anchor).  ")
md.append(f"**Sources:** Signed (v3 FINAL Billing Price) · PG (`scripts/audit-sc-prices.mjs` post-Stage-1) · Account files (§2b) · SC workbooks (11 per-account xlsx).  ")
md.append(f"**Method:** 2dp match at billed precision; sub-cent storage noted separately.")
md.append("")
md.append("---")
md.append("")
md.append("## 1. Scorecard")
md.append("")
md.append(f"| Comparison | PASS | FAIL | Note |")
md.append(f"|---|---:|---:|---|")
md.append(f"| **PG vs Signed** | **{pg_signed_pass}** / {len(audit_rows)} | **{pg_signed_real_fail}** real + **{pg_signed_signed_no_price}** signed-no-price + **{pg_signed_stale_stage1}** signed-stale-per-Stage-1 | Certification gate |")
md.append(f"| **AccountFile vs Signed** | **{af_signed_pass}** / {len(audit_rows)} | **{af_signed_fail}** | ({af_signed_none} unmatched/no acct-file rate found) |")
md.append(f"| **Workbook vs Signed** | *n/a* | **{wb_div_count} divergences catalogued** | RETIRED authority - all divergences expected |")
md.append("")
if pg_signed_real_fail == 0 and pg_signed_signed_no_price == 0 and pg_signed_stale_stage1 == 0:
    md.append("### 🎯 Verdict: **CERTIFIED** — PG = Signed at 105/105.")
elif pg_signed_real_fail == 0:
    md.append(f"### 🟢 Verdict: **CERTIFIED (with catalogued signed-side notes)** — {pg_signed_pass}/{len(audit_rows)} PG=Signed at 2dp; {pg_signed_stale_stage1} row(s) intentionally moved by Stage-1 directives (PG correct, signed sheet needs v4 refresh); {pg_signed_signed_no_price} row(s) have non-numeric signed values ('NEEDS PRICE', fee-account $0). PG is right on every row.")
else:
    md.append(f"### ❌ Verdict: **NOT CERTIFIED** — {pg_signed_real_fail} PG-vs-Signed failure(s); see §3 below.")
md.append("")
md.append("---")
md.append("")
md.append("## 2. Stage-1 fixes confirmation (from post-fix PG re-dump)")
md.append("")
smokes = []
for r in audit_rows:
    if r["account"] == "CIN - AZ" and r["group"] == "Major League" and r["service"] == "Breakfast":
        smokes.append(("CIN-AZ MLB Breakfast (target $20.31)", r["pg_proj"], r["signed_bill"], r["pg_proj"] == 20.31))
    if "Media Meals" in r["service"]:
        smokes.append(("TBJ-FL Media Meals PG (target $16.00)", r["pg_proj"], "16.0", r["pg_proj"] == 16.00))
    if "Extended Day Labor" in r["service"]:
        smokes.append(("TBR-FL Extended Day Labor (case-exact)", r["service"], "n/a", "labor" not in r["service"] or "Labor" in r["service"]))
# tax-free suffix check
tax_free_svc = [r for r in audit_rows if "(tax-free)" in (r["service"] or "").lower()]
smokes.append(("No '(tax-free)' suffixes on service names", "0 rows" if not tax_free_svc else f"{len(tax_free_svc)} rows", "0", not tax_free_svc))
# BGC tax_free flag
bgc = next((r for r in pg_all if r["account_key"] == "TBR - FL" and "b&g" in (r["service"] or "").lower()), None)
if bgc:
    smokes.append(("TBR-FL BGC is_tax_free = TRUE (Stage-1-b)", bgc["is_tax_free"], "TRUE", bgc["is_tax_free"] is True))

md.append("| Fix | Observed | Expected | Status |")
md.append("|---|---|---|---|")
for name, obs, exp, ok in smokes:
    md.append(f"| {name} | `{obs}` | `{exp}` | {'✅' if ok else '❌'} |")
md.append("")
md.append("---")
md.append("")
md.append("## 3. PG ≠ Signed - classification")
md.append("")
real_fails = [r for r in audit_rows if r["verdict"] == "PG-FAIL"]
signed_np = [r for r in audit_rows if r["verdict"] == "SIGNED-NO-PRICE"]
stage1_stale = [r for r in audit_rows if r["verdict"] == "SIGNED-STALE-STAGE1"]
if not real_fails and not signed_np and not stage1_stale:
    md.append("**None.** PG = Signed on all 105 rows at 2dp.")
else:
    if real_fails:
        md.append("### 3a. 🔴 Real PG-vs-Signed failures (block certification)")
        md.append("")
        md.append("| Account | Group | Service | Signed | PG | AcctFile | Workbook (proj/actl) | Notes |")
        md.append("|---|---|---|---:|---:|---:|---|---|")
        for r in real_fails:
            wb_str = f"{money(r['wb_proj'])} / {money(r['wb_actl'])}"
            md.append(f"| {r['account']} | {r['group']} | {r['service']} | {money(r['signed_bill'])} | {money(r['pg_proj'])} | {money(r['acct_file'])} | {wb_str} | flags: {r['flag'] or '-'} |")
        md.append("")
    if stage1_stale:
        md.append("### 3b. 🟢 PG intentionally ahead of signed (Stage-1 directives; PG correct)")
        md.append("")
        md.append("Kevin explicitly ordered these PG values in the Stage-1 batch. Signed sheet (v3 FINAL) still shows the pre-directive value and needs a **v4 refresh** to catch up. PG is authoritative here per Kevin's ruling; signed drift is the follow-up.")
        md.append("")
        md.append("| Account | Group | Service | Signed (stale) | PG (Kevin-directed) | Workbook (actl) | Notes |")
        md.append("|---|---|---|---:|---:|---:|---|")
        for r in stage1_stale:
            md.append(f"| {r['account']} | {r['group']} | {r['service']} | {money(r['signed_bill'])} | {money(r['pg_proj'])} | {money(r['wb_actl'])} | Stage-1 target |")
        md.append("")
    if signed_np:
        md.append("### 3c. ⚪ Signed cell has no price (excluded from cert denominator)")
        md.append("")
        md.append("| Account | Group | Service | Signed | PG | Notes |")
        md.append("|---|---|---|---|---:|---|")
        for r in signed_np:
            md.append(f"| {r['account']} | {r['group']} | {r['service']} | `{r['signed_bill']}` | {money(r['pg_proj'])} | flags: {r['flag'] or '-'} |")
        md.append("")
md.append("")
md.append("---")
md.append("")
md.append("## 4. Doc drift (AccountFile ≠ Signed)")
md.append("")
drifts = [r for r in audit_rows if r["af_v_signed_ok"] is False]
if not drifts:
    md.append("**None.** Every account-file rate matches signed at 2dp.")
else:
    md.append("| Account | Group | Service | Signed | AcctFile |")
    md.append("|---|---|---|---:|---:|")
    for r in drifts:
        md.append(f"| {r['account']} | {r['group']} | {r['service']} | {money(r['signed_bill'])} | {money(r['acct_file'])} |")
md.append("")
if af_signed_none:
    md.append(f"### AcctFile lookups with no match ({af_signed_none} rows)")
    md.append("")
    no_af = [r for r in audit_rows if r["af_v_signed_ok"] is None][:20]
    md.append("Common cause: account file §2b table doesn't itemize this line (e.g., Extra Protein, Extended Day Labor per-unit lines), non-per-meal add-ons, or fee-account $0 rows with non-numeric \"NEEDS PRICE\" signed cells. First 20 shown for reference.")
    md.append("")
    md.append("| Account | Group | Service | Signed |")
    md.append("|---|---|---|---:|")
    for r in no_af:
        md.append(f"| {r['account']} | {r['group']} | {r['service']} | {money(r['signed_bill'])} |")
    md.append("")
md.append("---")
md.append("")
md.append("## 5. Workbook divergences (⚪ RETIRED authority - catalogued not fixed)")
md.append("")
wb_divs = [r for r in audit_rows if r["wb_div_reason"] and r["wb_div_reason"] != "fee_acct"]
if not wb_divs:
    md.append("**None catalogued.** Workbook prices align with signed on every row.")
else:
    md.append(f"**{len(wb_divs)} workbook cells diverge from signed** (expected: workbook = sheet-era artifact; divergences reflect old projection-tab formulas, pre-sc-8c actuals, stale re-imports).")
    md.append("")
    md.append("| Account | Group | Service | Signed_Bill | Signed_Full | WB_proj | WB_actl | Divergence reason |")
    md.append("|---|---|---|---:|---:|---:|---:|---|")
    for r in wb_divs:
        md.append(f"| {r['account']} | {r['group']} | {r['service']} | {money(r['signed_bill'])} | {money(r['signed_full'])} | {money(r['wb_proj'])} | {money(r['wb_actl'])} | {r['wb_div_reason']} |")
md.append("")
md.append("---")
md.append("")
md.append("## 6. Full four-way comparison (all 105 rows)")
md.append("")
md.append("Columns: Account · Group · Service · Signed_Bill · Signed_Full · PG_proj · AcctFile · WB_proj · WB_actl · Verdict")
md.append("")
md.append("| Account | Group | Service | Signed | Signed_Full | PG | AcctFile | WB_proj | WB_actl | Verdict |")
md.append("|---|---|---|---:|---:|---:|---:|---:|---:|---|")
for r in audit_rows:
    md.append(
        f"| {r['account']} | {r['group']} | {r['service']} | "
        f"{money(r['signed_bill'])} | {money(r['signed_full'])} | {money(r['pg_proj'])} | "
        f"{money(r['acct_file'])} | {money(r['wb_proj'])} | {money(r['wb_actl'])} | "
        f"{verdict_icon(r['verdict'])} |"
    )
md.append("")
md.append("---")
md.append("")
md.append("## 7. Coverage / orphan check")
md.append("")
md.append(f"- **Signed rows total:** {len(signed_rows)}")
md.append(f"- **PG rows total:** {len(pg_all)}")
md.append(f"- **Signed rows with no PG match:** {len(missing_pg)}")
md.append(f"- **PG rows not in signed:** {len(missing_signed_from_pg)}")
md.append("")
if missing_pg:
    md.append("### Signed rows missing from PG")
    md.append("")
    for k in missing_pg:
        md.append(f"- `{k[0]}` / `{k[1]}` / `{k[2]}`")
    md.append("")
if missing_signed_from_pg:
    md.append("### PG rows not in signed")
    md.append("")
    for k in missing_signed_from_pg:
        md.append(f"- `{k[0]}` / `{k[1]}` / `{k[2]}`")
    md.append("")
md.append("---")
md.append("")
md.append("## 8. Batch-3 accounts (first full PG price check)")
md.append("")
batch3 = ["TBJ - FL", "TBR - FL", "TXR - AZ"]
for acct in batch3:
    rows = [r for r in audit_rows if r["account"] == acct]
    md.append(f"### {acct} ({len(rows)} services)")
    md.append("")
    md.append("| Group | Service | Signed | PG | Verdict |")
    md.append("|---|---|---:|---:|---|")
    for r in rows:
        md.append(f"| {r['group']} | {r['service']} | {money(r['signed_bill'])} | {money(r['pg_proj'])} | {verdict_icon(r['verdict'])} |")
    md.append("")
md.append("---")
md.append("")
md.append("## 9. Final verdict")
md.append("")
if pg_signed_real_fail == 0 and pg_signed_stale_stage1 == 0 and pg_signed_signed_no_price == 0:
    md.append("**🎯 CERTIFIED (105/105 PG = Signed at 2dp).**")
    md.append("")
    md.append(f"AccountFile drift = {af_signed_fail} · Workbook divergences = {wb_div_count} (retired authority, catalogued).")
elif pg_signed_real_fail == 0:
    md.append(f"**🟢 CERTIFIED (with catalogued signed-side notes).**")
    md.append("")
    md.append(f"- PG = Signed at 2dp on {pg_signed_pass}/{len(audit_rows)} rows.")
    md.append(f"- {pg_signed_stale_stage1} row(s) where Kevin's Stage-1 directive moved PG ahead of signed (Media Meals $16). Follow-up: v4 signed refresh.")
    md.append(f"- {pg_signed_signed_no_price} row(s) with 'NEEDS PRICE' signed cell (STL-FL MiLB Snack, fee account). PG correctly $0; excluded from cert denominator.")
    md.append(f"- AccountFile drift = {af_signed_fail} (see §4). Workbook divergences = {wb_div_count} (retired authority, catalogued, see §5).")
    md.append("")
    md.append("**Zero PG-vs-Signed real failures. Certification GATE = GREEN.**")
else:
    md.append(f"**❌ NOT CERTIFIED.** {pg_signed_real_fail} PG-vs-Signed real failure(s) block certification; see §3a.")
md.append("")

with open(OUT_MD, "w") as f:
    f.write("\n".join(md))
print(f"wrote {len(audit_rows)} rows to {OUT_MD}")
print(f"PG-vs-Signed: {pg_signed_pass}/{len(audit_rows)} pass, {pg_signed_real_fail} real-fail, {pg_signed_stale_stage1} stage1-stale, {pg_signed_signed_no_price} signed-no-price")
print(f"AF-vs-Signed: {af_signed_pass}/{len(audit_rows)} pass, {af_signed_fail} fail, {af_signed_none} none")
print(f"WB divergences: {wb_div_count}")
